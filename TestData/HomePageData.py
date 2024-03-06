import openpyxl


class HomePageData:
    test_HomePageData = [{"firstname": "Tester", "email": "testing@gmail.com", "password": "abcdefg", "gender": "Male"},
                         {"firstname": "Tester1", "email": "testing1@gmail.com", "password": "abcd1234",
                          "gender": "Male"}]

    @staticmethod
    def getTestData(testcaseName):
        book = openpyxl.load_workbook("C:\\ExcelDataSelenium\\formaData.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcaseName:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
